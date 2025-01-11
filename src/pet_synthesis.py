import os
import logging
import itertools
import numpy as np
import torch
import pickle
from torch.utils.data import DataLoader
import torch.nn.functional as F
from tqdm import tqdm
from sklearn.metrics import roc_curve, auc, accuracy_score
from sklearn.metrics import confusion_matrix, ConfusionMatrixDisplay
import matplotlib

matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

from .utils import *
from .dataset import PETDataset, InfiniteSamplerWrapper, load_and_resize_image
from .loss import GANLoss


def train_generator(config):
    # Loggings
    open_log(config.LOG_PATH)

    # ----------------------------------------
    #      Initialize training parameters
    # ----------------------------------------

    # Build generators
    if config.MODE == 'train' and config.PRETRAIN_GEN:
        netG_name1 = os.path.join(config.CKPT_PATH, 'gen_to_CFT_{:d}.pth'.format(config.PRETRAIN_GEN))
        netG_name2 = os.path.join(config.CKPT_PATH, 'gen_to_FDG_{:d}.pth'.format(config.PRETRAIN_GEN))
    else:
        netG_name1 = None
        netG_name2 = None
    netG_to_CFT = create_generator(config, netG_name1).cuda()
    netG_to_FDG = create_generator(config, netG_name2).cuda()

    # Build discriminators
    if config.MODE == 'train' and config.PRETRAIN_DIS:
        netD_name1 = os.path.join(config.CKPT_PATH, 'dis_to_CFT_{:d}.pth'.format(config.PRETRAIN_DIS))
        netD_name2 = os.path.join(config.CKPT_PATH, 'dis_to_FDG_{:d}.pth'.format(config.PRETRAIN_DIS))
    else:
        netD_name1 = None
        netD_name2 = None
    netD_FDG = create_discriminator(config, netD_name1).cuda()
    netD_CFT = create_discriminator(config, netD_name2).cuda()

    # Optimizers
    optimizer_netG = torch.optim.Adam(itertools.chain(netG_to_CFT.parameters(), netG_to_FDG.parameters()),
                                      lr=config.LR_G, betas=(config.BETA1, config.BETA2))
    optimizer_netD = torch.optim.Adam(itertools.chain(netD_CFT.parameters(), netD_FDG.parameters()),
                                      lr=config.LR_D, betas=(config.BETA1, config.BETA2))

    # Loss function
    GAN_criterion = GANLoss(gan_mode=config.GAN_MODE).cuda()

    # ----------------------------------------
    #       Initialize training dataset
    # ----------------------------------------

    # Define training dataset
    train_data = PETDataset(config.DATAROOT, config.TRAIN_DATA, config.IMAGE_SIZE, config.DATA_NORM)
    logging.info('The overall number of training images is %d' % len(train_data))

    # Define training dataloader
    train_loader = iter(DataLoader(train_data, batch_size=config.BATCH_SIZE, drop_last=True,
                                   sampler=InfiniteSamplerWrapper(train_data),
                                   num_workers=config.NUM_WORKERS))

    # ----------------------------------------
    #      Initialize validation dataset
    # ----------------------------------------

    # Define validation dataset
    val_data = PETDataset(config.DATAROOT, config.TEST_DATA, config.IMAGE_SIZE, config.DATA_NORM)
    logging.info('The overall number of validation images is %d' % len(val_data))

    # Define validation dataloader
    val_loader = DataLoader(val_data, batch_size=1, num_workers=config.NUM_WORKERS)
    sample_iterator = val_data.create_iterator(config.VIS_NUM)

    # ----------------------------------------
    #              Start Training
    # ----------------------------------------

    # Load pretrained models
    if config.PRETRAIN_GEN:
        start_iter = config.PRETRAIN_GEN
        logging.info("Resume training from iteration %d" % start_iter)
    else:
        start_iter = 0
        logging.info("Start training ......")

    # Training loop
    for iter_idx in tqdm(range(start_iter, config.TOTAL_ITERS), desc="Training"):
        cur_iter = iter_idx + 1  # current iteration

        # Load images ([B, 1, D, H, W])
        data = next(train_loader)
        real_FDG_img, real_CFT_img = data["FDG_img"].cuda().float(), data["CFT_img"].cuda().float()
        label = data["label"].cuda().float()
        paired_flag = data["paired_flag"]

        # ----------------------------------------
        #            Train Generators
        # ----------------------------------------

        # FDG --> CFT --> FDG
        fake_CFT_img = netG_to_CFT(real_FDG_img)
        recon_FDG_img = netG_to_FDG(fake_CFT_img)

        # CFT --> FDG --> CFT
        fake_FDG_img = netG_to_FDG(real_CFT_img)
        recon_CFT_img = netG_to_CFT(fake_FDG_img)

        # (1) compute cycle loss
        cycle_loss_FDG = F.l1_loss(recon_FDG_img, real_FDG_img) * config.CYCLE_WEIGHT
        cycle_loss_CFT = F.l1_loss(recon_CFT_img, real_CFT_img) * config.CYCLE_WEIGHT
        cycle_loss = cycle_loss_FDG + cycle_loss_CFT

        # (2) compute GAN loss
        if config.GAN_START_ITER and cur_iter >= config.GAN_START_ITER:
            fake_CFT_out = netD_CFT(fake_CFT_img, label)
            fake_FDG_out = netD_FDG(fake_FDG_img, label)
            GAN_loss_G = (GAN_criterion(fake_CFT_out, True, for_discriminator=False) +
                          GAN_criterion(fake_FDG_out, True, for_discriminator=False)) * config.GAN_WEIGHT
        else:
            GAN_loss_G = 0

        # (3) compute paired reconstruction loss
        recon_loss_CFT = 0
        recon_loss_FDG = 0
        paired_counts = 0
        for batch_idx, is_paired in enumerate(paired_flag):
            if is_paired:
                recon_loss_CFT += F.l1_loss(fake_CFT_img[batch_idx], real_CFT_img[batch_idx])
                recon_loss_FDG += F.l1_loss(fake_FDG_img[batch_idx], real_FDG_img[batch_idx])
                paired_counts += 1
        if paired_counts > 0:
            recon_loss_CFT = recon_loss_CFT / paired_counts * config.RECON_WEIGHT
            recon_loss_FDG = recon_loss_FDG / paired_counts * config.RECON_WEIGHT
        recon_loss = recon_loss_CFT + recon_loss_FDG

        # backward and optimize
        total_loss_G = cycle_loss + GAN_loss_G + recon_loss
        optimizer_netG.zero_grad()
        total_loss_G.backward()
        optimizer_netG.step()

        # ----------------------------------------
        #          Train Discriminators
        # ----------------------------------------

        if config.GAN_START_ITER and cur_iter >= config.GAN_START_ITER:
            # (1) compute loss with real images
            real_FDG_out = netD_FDG(real_FDG_img, label)
            real_CFT_out = netD_CFT(real_CFT_img, label)
            GAN_loss_D_real = GAN_criterion(real_FDG_out, True, for_discriminator=True) + GAN_criterion(real_CFT_out, True, for_discriminator=True)

            # (2) compute loss with fake images
            with torch.no_grad():
                fake_CFT_img = netG_to_CFT(real_FDG_img, label)
                fake_FDG_img = netG_to_FDG(real_CFT_img, label)
            fake_CFT_out = netD_CFT(fake_CFT_img, label)
            fake_FDG_out = netD_FDG(fake_FDG_img, label)
            GAN_loss_D_fake = GAN_criterion(fake_CFT_out, False, for_discriminator=True) + GAN_criterion(fake_FDG_out, False, for_discriminator=True)

            # (3) backward and optimize
            GAN_loss_D = (GAN_loss_D_real + GAN_loss_D_fake) * 0.5 * config.GAN_WEIGHT
            optimizer_netD.zero_grad()
            GAN_loss_D.backward()
            optimizer_netD.step()

        # ----------------------------------------
        #            Log training states
        # ----------------------------------------

        if config.LOG_INTERVAL and cur_iter % config.LOG_INTERVAL == 0:
            # print training status for reconstruction
            if paired_counts > 0:
                logging.info('[Iteration {:d}] CFT Reconstruction Loss: {:.4f}'.format(cur_iter, recon_loss_CFT.item()))
                logging.info('[Iteration {:d}] FDG Reconstruction Loss: {:.4f}'.format(cur_iter, recon_loss_FDG.item()))

            # print training state for cycle consistency
            train_state_cycle_1 = '[Iteration {:d} | FDG -> CFT -> FDG] Cycle Loss: {:.4f}'.format(cur_iter,
                                                                                                   cycle_loss_FDG.item())
            train_state_cycle_2 = '[Iteration {:d} | CFT -> FDG -> CFT] Cycle Loss: {:.4f}'.format(cur_iter,
                                                                                                   cycle_loss_CFT.item())
            if config.GAN_START_ITER and cur_iter >= config.GAN_START_ITER:
                with torch.no_grad():
                    # (1) FDG --> CFT --> FDG
                    real_CFT_img_out = netD_CFT(real_CFT_img, label)
                    fake_CFT_img_out = netD_CFT(fake_CFT_img, label)
                    real_CFT_score = real_CFT_img_out.mean().item()
                    fake_CFT_score = fake_CFT_img_out.mean().item()

                    # (2) CFT --> FDG --> CFT
                    real_FDG_img_out = netD_FDG(real_FDG_img, label)
                    fake_FDG_img_out = netD_FDG(fake_FDG_img, label)
                    real_FDG_score = real_FDG_img_out.mean().item()
                    fake_FDG_score = fake_FDG_img_out.mean().item()
                train_state_cycle_1 += ' | D(y) / D(G(x)): {:.4f} / {:.4f}'.format(real_CFT_score, fake_CFT_score)
                train_state_cycle_2 += ' | D(x) / D(G(y)): {:.4f} / {:.4f}'.format(real_FDG_score, fake_FDG_score)
            logging.info(train_state_cycle_1)
            logging.info(train_state_cycle_2)
            logging.info("\n")

        # ----------------------------------------
        #               Validation
        # ----------------------------------------
        if (cur_iter % config.SAVE_INTERVAL == 0 and cur_iter >= config.SAVE_START) or cur_iter == config.TOTAL_ITERS:
            # 1. save models
            # (1) save generators
            save_model(netG_to_CFT, 'gen_to_CFT_{:d}.pth'.format(cur_iter), config.CKPT_PATH)
            save_model(netG_to_FDG, 'gen_to_FDG_{:d}.pth'.format(cur_iter), config.CKPT_PATH)
            logging.info('The generators are successfully saved at iteration {:d}'.format(cur_iter))
            # (2) Save discriminators
            if config.GAN_START_ITER and cur_iter >= config.GAN_START_ITER:
                save_model(netD_CFT, 'dis_to_CFT_{:d}.pth'.format(cur_iter), config.CKPT_PATH)
                save_model(netD_FDG, 'dis_to_FDG_{:d}.pth'.format(cur_iter), config.CKPT_PATH)
                logging.info('The discriminators are successfully saved at iteration {:d}'.format(cur_iter))

            # 2. Calculate metrics
            val_metrics = {}
            for data_type in ["paired", "unpaired"]:
                val_metrics[f"{data_type}_psnr_FDG"] = []
                val_metrics[f"{data_type}_ssim_FDG"] = []
                val_metrics[f"{data_type}_psnr_CFT"] = []
                val_metrics[f"{data_type}_ssim_CFT"] = []

            with torch.no_grad():
                for _, data in enumerate(val_loader):
                    # (1) Load images ([B, 1, D, H, W])
                    real_FDG_img, real_CFT_img = data["FDG_img"].cuda().float(), data["CFT_img"].cuda().float()
                    is_paired = data["paired_flag"][0]

                    # (2) paired data results
                    if is_paired:
                        syn_CFT_img = netG_to_CFT(real_FDG_img)
                        syn_FDG_img = netG_to_FDG(real_CFT_img)
                        val_metrics["paired_psnr_FDG"].append(compute_psnr(syn_FDG_img, real_FDG_img, config.DATA_NORM))
                        val_metrics["paired_ssim_FDG"].append(compute_ssim(syn_FDG_img, real_FDG_img, config.DATA_NORM))
                        val_metrics["paired_psnr_CFT"].append(compute_psnr(syn_CFT_img, real_CFT_img, config.DATA_NORM))
                        val_metrics["paired_ssim_CFT"].append(compute_ssim(syn_CFT_img, real_CFT_img, config.DATA_NORM))

                    # (3) unpaired data results
                    else:
                        # FDG -> CFT -> FDG
                        fake_CFT_img = netG_to_CFT(real_FDG_img)
                        recon_FDG_img = netG_to_FDG(fake_CFT_img)
                        # CFT -> FDG -> CFT
                        fake_FDG_img = netG_to_FDG(real_CFT_img)
                        recon_CFT_img = netG_to_CFT(fake_FDG_img)
                        val_metrics["unpaired_psnr_FDG"].append(
                            compute_psnr(recon_FDG_img, real_FDG_img, config.DATA_NORM))
                        val_metrics["unpaired_ssim_FDG"].append(
                            compute_ssim(recon_FDG_img, real_FDG_img, config.DATA_NORM))
                        val_metrics["unpaired_psnr_CFT"].append(
                            compute_psnr(recon_CFT_img, real_CFT_img, config.DATA_NORM))
                        val_metrics["unpaired_ssim_CFT"].append(
                            compute_ssim(recon_CFT_img, real_CFT_img, config.DATA_NORM))

            # (4) Log metrics for paired data
            logging.info(
                '[Iteration {:d} - Validation | Synthetic CFT] PSNR: {:.4f} ± {:.4f} | SSIM: {:.4f} ± {:.4f}'
                .format(cur_iter, np.mean(val_metrics["paired_psnr_CFT"]), np.std(val_metrics["paired_psnr_CFT"]),
                        np.mean(val_metrics["paired_ssim_CFT"]), np.std(val_metrics["paired_ssim_CFT"])))
            logging.info(
                '[Iteration {:d} - Validation | Synthetic FDG] PSNR: {:.4f} ± {:.4f} | SSIM: {:.4f} ± {:.4f}'
                .format(cur_iter, np.mean(val_metrics["paired_psnr_FDG"]), np.std(val_metrics["paired_psnr_FDG"]),
                        np.mean(val_metrics["paired_ssim_FDG"]), np.std(val_metrics["paired_ssim_FDG"])))

            # (5) Log metrics for unpaired data
            logging.info(
                '[Iteration {:d} - Validation | FDG -> CFT -> FDG] PSNR: {:.4f} ± {:.4f} | SSIM: {:.4f} ± {:.4f}'
                .format(cur_iter, np.mean(val_metrics["unpaired_psnr_FDG"]), np.std(val_metrics["unpaired_psnr_FDG"]),
                        np.mean(val_metrics["unpaired_ssim_FDG"]), np.std(val_metrics["unpaired_ssim_FDG"])))
            logging.info(
                '[Iteration {:d} - Validation | CFT -> FDG -> CFT] PSNR: {:.4f} ± {:.4f} | SSIM: {:.4f} ± {:.4f}'
                .format(cur_iter, np.mean(val_metrics["unpaired_psnr_CFT"]), np.std(val_metrics["unpaired_psnr_CFT"]),
                        np.mean(val_metrics["unpaired_ssim_CFT"]), np.std(val_metrics["unpaired_psnr_CFT"])))

            # 3. Visualization (FDG -> CFT -> FDG)
            # (1) input
            data = next(sample_iterator)
            real_FDG_img = data["FDG_img"].cuda().float()
            # (2) output
            with torch.no_grad():
                fake_CFT_img = netG_to_CFT(real_FDG_img)
                recon_FDG_img = netG_to_FDG(fake_CFT_img)
            error_map = get_error_map(recon_FDG_img, real_FDG_img)
            # (3) save
            img_list = [torch.flip(real_FDG_img[:, :, config.VIS_SLICE - 1], dims=[2]),
                        torch.flip(fake_CFT_img[:, :, config.VIS_SLICE - 1], dims=[2]),
                        torch.flip(recon_FDG_img[:, :, config.VIS_SLICE - 1], dims=[2]),
                        torch.flip(error_map[:, :, config.VIS_SLICE - 1], dims=[2])]
            filename = "iteration_{:d}.jpg".format(cur_iter)
            show_image(img_list, filename, config.SAMPLE_PATH, config.DATA_NORM, config.VIS_NUM)

    print("Training finished !!")


def test_generator(config):
    # Loggings
    result_path = os.path.join(config.TEST_PATH, str(config.MODEL_GEN))
    os.makedirs(result_path, exist_ok=True)

    # ----------------------------------------
    #      Initialize testing parameters
    # ----------------------------------------

    # Build network
    if config.MODE == 'test' and config.MODEL_GEN:
        netG_name1 = os.path.join(config.CKPT_PATH, 'gen_to_CFT_{:d}.pth'.format(config.MODEL_GEN))
        netG_name2 = os.path.join(config.CKPT_PATH, 'gen_to_FDG_{:d}.pth'.format(config.MODEL_GEN))
    else:
        netG_name1 = None
        netG_name2 = None
    netG_to_CFT = create_generator(config, netG_name1).cuda()
    netG_to_FDG = create_generator(config, netG_name2).cuda()
    netG_to_CFT.eval()
    netG_to_FDG.eval()

    # ----------------------------------------
    #       Initialize testing dataset
    # ----------------------------------------

    # Define testing dataset
    test_data = PETDataset(config.DATAROOT, config.TEST_DATA, config.IMAGE_SIZE, config.DATA_NORM)
    print('The overall number of testing images is %d\n' % len(test_data))

    # Define testing dataloader
    test_loader = DataLoader(test_data, batch_size=1, shuffle=False, num_workers=config.NUM_WORKERS)

    # ----------------------------------------
    #              Load Templates
    # ----------------------------------------
    mask_root = os.path.join(config.DATAROOT, "templates")

    # (1) reference ROI
    cerebelGM_mask = load_and_resize_image(os.path.join(mask_root, "cerebellumGM.nii"))

    # (2) target ROI
    # anterior putamen
    a_putamen_left_mask = load_and_resize_image(os.path.join(mask_root, "a-putamen_l.nii"))
    a_putamen_right_mask = load_and_resize_image(os.path.join(mask_root, "a-putamen_r.nii"))
    # posterior putament
    p_putamen_left_mask = load_and_resize_image(os.path.join(mask_root, "p-putamen_l.nii"))
    p_putamen_right_mask = load_and_resize_image(os.path.join(mask_root, "p-putamen_r.nii"))
    # caudate
    caudate_left_mask = load_and_resize_image(os.path.join(mask_root, "CAU_l.nii"))
    caudate_right_mask = load_and_resize_image(os.path.join(mask_root, "CAU_r.nii"))

    # (3) collect all ROI regions
    syn_roi_regions = {
        "aPUT_l": a_putamen_left_mask, "aPUT_r": a_putamen_right_mask,
        "pPUT_l": p_putamen_left_mask, "pPUT_r": p_putamen_right_mask,
        "CAU_l": caudate_left_mask, "CAU_r": caudate_right_mask,
    }

    sbr_roi_regions = {
        "aPUT": a_putamen_left_mask + a_putamen_right_mask,
        "pPUT": p_putamen_left_mask + p_putamen_right_mask,
        "CAU": caudate_left_mask + caudate_right_mask,
    }

    # ----------------------------------------
    #                Metrics
    # ----------------------------------------

    syn_results = {}
    for category in ["PD", "NC", "total"]:
        syn_results[category] = {"paired_img_name": []}
        for roi in syn_roi_regions.keys():
            for metric in ["psnr", "ssim"]:
                syn_results[category].update({f"{roi}_{metric}": []})

    sbr_results = {"img_name": [], "paired": [], "y_true": []}
    for roi in sbr_roi_regions.keys():
        sbr_results[roi] = {"y_score": [], "auc": -1, "thresh": -1,
                            "y_pred": [], "accuracy": -1, "sensitivity": -1, "specificity": -1, "F1 score": -1}

    # ----------------------------------------
    #              Start Testing
    # ----------------------------------------

    print("Start testing ......")
    with torch.no_grad():
        for _, data in tqdm(enumerate(test_loader), total=len(test_loader), desc="Testing"):

            # Load images ([B, 1, D, H, W])
            real_FDG_img, real_CFT_img = data["FDG_img"].cuda().float(), data["CFT_img"].cuda().float()
            label = data["label"].cuda().float()
            is_paired = data["paired_flag"][0]
            save_img_name = data["img_name"][0]
            category = save_img_name.split("_")[0]

            # inference
            fake_CFT_img = netG_to_CFT(real_FDG_img)

            # save visualization results
            if config.SAVE_RESULTS:
                save_vol_image(real_FDG_img, save_img_name, os.path.join(result_path, "input_FDG"), config.DATA_NORM)
                save_vol_image(fake_CFT_img, save_img_name, os.path.join(result_path, "syn_CFT"), config.DATA_NORM)
                if is_paired:
                    save_vol_image(real_CFT_img, save_img_name, os.path.join(result_path, "real_CFT"), config.DATA_NORM)
                    save_vol_image(get_error_map(fake_CFT_img, real_CFT_img), save_img_name,
                                   os.path.join(result_path, "syn_CFT_error_map"), config.DATA_NORM)

            # synthesis metrics
            if is_paired:
                syn_results["total"]["paired_img_name"].append(save_img_name)
                syn_results[category]["paired_img_name"].append(save_img_name)
                for roi in syn_roi_regions.keys():
                    # (1) compute all metrcis
                    psnr = compute_psnr(fake_CFT_img, real_CFT_img, config.DATA_NORM, syn_roi_regions[roi])
                    ssim = compute_ssim(fake_CFT_img, real_CFT_img, config.DATA_NORM, syn_roi_regions[roi])
                    # (2) collect all metrics
                    # total results
                    syn_results["total"][f"{roi}_psnr"].append(psnr)
                    syn_results["total"][f"{roi}_ssim"].append(ssim)
                    # results of each category
                    syn_results[category][f"{roi}_psnr"].append(psnr)
                    syn_results[category][f"{roi}_ssim"].append(ssim)

            # SBR results
            sbr_results["img_name"].append(save_img_name)
            sbr_results["y_true"].append(label[:, 1].item())
            for roi, roi_mask in sbr_roi_regions.items():
                sbr = compute_SBR(fake_CFT_img, cerebelGM_mask, config.DATA_NORM, roi_mask)
                sbr_results[roi]["y_score"].append(sbr)

    # ----------------------------------------
    #               save results
    # ----------------------------------------

    # 1. synthesis results
    overall_df = {}
    for category in ["PD", "NC", "total"]:
        overall_df[category] = pd.DataFrame({"ROI": syn_roi_regions.keys()},
                                            index=range(1, len(syn_roi_regions.keys()) + 1))
        for metric in ["psnr", "ssim"]:
            column_values = []
            for roi in syn_roi_regions.keys():
                metric_name = f"{roi}_{metric}"
                column_values.append("{:.4f} ± {:.4f}".format(
                    np.mean(syn_results[category][metric_name]), np.std(syn_results[category][metric_name]),
                ))
            overall_df[category][metric] = column_values

    xls_file = os.path.join(result_path, 'synthesis_results.xlsx')
    with pd.ExcelWriter(xls_file, engine='openpyxl') as writer:
        for sheet_name in ["PD", "NC", "total"]:
            overall_df[sheet_name].to_excel(writer, sheet_name=sheet_name, index=False)
            worksheet = writer.sheets[sheet_name]
            for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row, min_col=1,
                                           max_col=worksheet.max_column):
                for cell in row:
                    if cell.row == 1:
                        cell.font = openpyxl.styles.Font(name='Times New Roman', size=14)
                    else:
                        cell.font = openpyxl.styles.Font(name='Times New Roman', size=12)

    # 2. SBR results
    for roi in sbr_roi_regions.keys():
        # (1) draw ROC curve
        inverted_scores = -np.array(sbr_results[roi]["y_score"])  #
        fpr, tpr, thresholds = roc_curve(sbr_results["y_true"], inverted_scores, pos_label=1.0)
        sbr_results[roi]["auc"] = auc(fpr, tpr)
        print('[{:s}] AUC: {:.3f}'.format(roi, sbr_results[roi]["auc"]))

        plt.figure()
        plt.plot(fpr, tpr, color='darkorange', lw=2, label='ROC curve (AUC = %0.4f)' % sbr_results[roi]["auc"])
        plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
        plt.xlim([-0.05, 1.0])
        plt.ylim([0.0, 1.05])
        plt.xlabel("False Positive Rate")
        plt.ylabel("True Positive Rate")
        plt.title(f"[{roi}] ROC Curve")
        plt.legend(loc="lower right")
        plt.savefig(os.path.join(result_path, f"{roi}_roc_curve.png"), format='png', dpi=300)

        # (2) classification
        youden_thresh = find_youden_index_threshold(fpr, tpr, thresholds)
        sbr_results[roi]["thresh"] = -youden_thresh
        sbr_results[roi]["y_pred"] = np.where(inverted_scores >= youden_thresh, 1, 0)
        cm = confusion_matrix(sbr_results["y_true"], sbr_results[roi]["y_pred"])
        sbr_results[roi]["accuracy"] = accuracy_score(sbr_results["y_true"], sbr_results[roi]["y_pred"])
        sbr_results[roi]["sensitivity"] = cm[1, 1] / (cm[1, 0] + cm[1, 1])
        sbr_results[roi]["specificity"] = cm[0, 0] / (cm[0, 0] + cm[0, 1])
        precision = cm[1, 1] / (cm[0, 1] + cm[1, 1])
        recall = sbr_results[roi]["sensitivity"]
        sbr_results[roi]["F1 score"] = 2 * precision * recall / (precision + recall)
        print(
            '[{:s} | SBR Thresh: {:.4f}] Accuracy: {:.3f} | Sensitivity: {:.3f} | Specificity: {:.3f} | F1 Score: {:.3f}\n'
            .format(roi, sbr_results[roi]["thresh"],
                    sbr_results[roi]["accuracy"],
                    sbr_results[roi]["sensitivity"],
                    sbr_results[roi]["specificity"],
                    sbr_results[roi]["F1 score"])
        )
        disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=["NC", "PD"])
        disp.plot(cmap="Blues")
        plt.title(f"({roi}) Confusion Matrix")
        plt.savefig(os.path.join(result_path, f"{roi}_cm.png"))

    # Record results
    record_file = os.path.join(result_path, "SBR_ROI_analysis.txt")
    with open(record_file, "w", encoding='utf-8') as f:
        for roi in sbr_roi_regions.keys():
            f.write(
                "[{:s} | SBR Thresh: {:.4f}] AUC: {:.3f} | Accuracy: {:.3f} | Sensitivity: {:.3f} | Specificity: {:.3f} | F1 Score: {:.3f}\n"
                .format(roi, sbr_results[roi]["thresh"],
                        sbr_results[roi]["auc"],
                        sbr_results[roi]["accuracy"],
                        sbr_results[roi]["sensitivity"],
                        sbr_results[roi]["specificity"],
                        sbr_results[roi]["F1 score"]),
            )

    pkl_file = os.path.join(sbr_results, "sbr_results.pkl")
    with open(pkl_file, "wb") as file:
        pickle.dump(sbr_results, file)

    print("Testing finished !!")
