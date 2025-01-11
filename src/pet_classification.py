import logging
import copy
from tqdm import tqdm
import torch
from torch.utils.data import DataLoader
from torch.optim.lr_scheduler import CosineAnnealingWarmRestarts
from sklearn.metrics import roc_curve, auc, accuracy_score, confusion_matrix, ConfusionMatrixDisplay
import matplotlib
matplotlib.use('Agg')
import matplotlib.pyplot as plt
import pandas as pd
import openpyxl

from .utils import *
from .dataset import PETDataset


def train_classifier(config):
    # Loggings
    open_log(config.LOG_PATH)

    # ----------------------------------------
    #      Initialize training parameters
    # ----------------------------------------

    # build classifier
    if config.MODE == 'train' and config.PRETRAIN_CLS:
        netC_name = os.path.join(config.CKPT_PATH, f'cls_{config.PRETRAIN_CLS}.pth')
    else:
        netC_name = None
    netC = create_classifier(config, netC_name).cuda()

    if config.IN_CHANNELS == 1:
        logging.info(f"Classifier Input: {config.INPUT_TYPE}")
    elif config.IN_CHANNELS == 2:
        logging.info("Classifier Input: real FDG + syn CFT")
    else:
        raise ValueError(f"{config.IN_CHANNELS} is not valid input channels")

    # build generators
    netG = create_generator(config, config.GEN_PATH).cuda()
    netG.eval()
    print("Load Generator from '{:s}'".format(config.GEN_PATH))

    # optimizer
    optimizer = torch.optim.Adam(
        netC.parameters(),
        lr=config.OPTIMIZER.lr_c,
        betas=(config.OPTIMIZER.beta1, config.OPTIMIZER.beta2),
    )

    # scheduler
    scheduler = CosineAnnealingWarmRestarts(
        optimizer,
        T_0=config.OPTIMIZER.t_initial,
        eta_min=config.OPTIMIZER.lr_min,
        verbose=True,
    )

    # Loss function
    CLS_criterion = torch.nn.CrossEntropyLoss(reduction="sum").cuda()

    # ----------------------------------------
    #       Initialize training dataset
    # ----------------------------------------

    # Define training dataset
    train_data = PETDataset(config.DATAROOT, config.TRAIN_DATA, config.IMAGE_SIZE, config.DATA_NORM)
    logging.info('The overall number of training images is %d' % len(train_data))

    # Define training dataloader
    train_loader = DataLoader(train_data, batch_size=config.BATCH_SIZE, drop_last=True, num_workers=config.NUM_WORKERS)

    # ----------------------------------------
    #      Initialize validation dataset
    # ----------------------------------------

    # Define validation dataset
    val_data = PETDataset(config.DATAROOT, config.VAL_DATA, config.IMAGE_SIZE, config.DATA_NORM)
    logging.info('The overall number of validation images is %d' % len(val_data))

    # Define validation dataloader
    val_loader = DataLoader(val_data, batch_size=1, num_workers=config.NUM_WORKERS)

    # ----------------------------------------
    #              Start Training
    # ----------------------------------------

    # Load pretrained models
    if config.PRETRAIN_CLS:
        start_epoch = config.PRETRAIN_CLS
        logging.info("Resume training from epoch %d" % start_epoch)
    else:
        start_epoch = 0
        logging.info("Start training ......")

    # Training loop
    for epoch in tqdm(range(start_epoch, config.TOTAL_EPOCH), desc="Training:"):
        cur_epoch = epoch + 1

        netC.train()
        for batch_idx, data in enumerate(train_loader):
            # Load images ([B, 1, D, H, W]) and class label ([B, CLASS_NUM])
            real_FDG_img = data["FDG_img"].cuda().float()
            label = data["label"].cuda().float()

            # Predict coarse synthetic CFT image
            if config.IN_CHANNELS > 1:
                with torch.no_grad():
                    input_CFT_img = netG(real_FDG_img)
                cls_inputs = torch.cat([real_FDG_img, input_CFT_img], dim=1)
            else:
                if config.INPUT_TYPE == "real_FDG":
                    cls_inputs = real_FDG_img
                elif config.INPUT_TYPE == "syn_CFT":
                    with torch.no_grad():
                        input_CFT_img = netG(real_FDG_img)
                    cls_inputs = input_CFT_img
                else:
                    raise ValueError(f"{config.INPUT_TYPE} is not valid input type")

            # ----------------------------------------
            #          Train  Classifier
            # ----------------------------------------

            # diagnose NC / PD
            cls_out = netC(cls_inputs)

            # compute classification loss
            cls_loss = CLS_criterion(cls_out, label)

            # backward and optimize
            optimizer.zero_grad()
            cls_loss.backward()
            optimizer.step()

        # adjust learning rate
        scheduler.step(epoch)
        logging.info("Epoch {:d}/{:d} - lr: {:.8f} | loss: {:.3f}"
                     .format(cur_epoch, config.TOTAL_EPOCH, optimizer.param_groups[0]['lr'], cls_loss.item()))

        # ----------------------------------------
        #               Validation
        # ----------------------------------------
        if config.LOG_INTERVAL and cur_epoch % config.LOG_INTERVAL == 0:
            val_metrics = {
                "y_true": [], "y_score": [], "auc": -1,  # labels and scores
                "y_pred": [], "accuracy": -1, "sensitivity": -1, "specificity": -1, "F1 score": -1,
            }

            netC.eval()
            with torch.no_grad():
                for _, data in enumerate(val_loader):

                    # (1) load images ([B, 1, D, H, W]) and class label ([B, CLASS_NUM])
                    real_FDG_img = data["FDG_img"].cuda().float()
                    label = data["label"].cuda().float()

                    # (2) construct inputs to classifier
                    if config.IN_CHANNELS > 1:
                        with torch.no_grad():
                            input_CFT_img = netG(real_FDG_img)
                        cls_inputs = torch.cat([real_FDG_img, input_CFT_img], dim=1)
                    else:
                        if config.INPUT_TYPE == "real_FDG":
                            cls_inputs = real_FDG_img
                        elif config.INPUT_TYPE == "syn_CFT":
                            with torch.no_grad():
                                input_CFT_img = netG(real_FDG_img)
                            cls_inputs = input_CFT_img
                        else:
                            raise ValueError(f"{config.INPUT_TYPE} is not valid input type")

                    # (3) classification
                    cls_out = netC(cls_inputs)
                    cls_pred_scores = torch.softmax(cls_out, dim=1)
                    val_metrics["y_true"].append(label[:, 1].item())
                    val_metrics["y_score"].append(cls_pred_scores[:, 1].item())

            # (4) calculate metrics
            fpr, tpr, thresholds = roc_curve(val_metrics["y_true"], val_metrics["y_score"])
            val_metrics["auc"] = auc(fpr, tpr)
            logging.info("Epoch {:d}/{:d} - AUC: {:.3f}".format(cur_epoch, config.TOTAL_EPOCH, val_metrics["auc"]))

            youden_thresh = find_youden_index_threshold(fpr, tpr, thresholds)
            val_metrics["y_pred"] = np.where(np.asarray(val_metrics["y_score"]) >= youden_thresh, 1, 0)
            cm = confusion_matrix(val_metrics["y_true"], val_metrics["y_pred"])
            val_metrics["accuracy"] = accuracy_score(val_metrics["y_true"], val_metrics["y_pred"])
            val_metrics["sensitivity"] = cm[1, 1] / (cm[1, 0] + cm[1, 1])
            val_metrics["specificity"] = cm[0, 0] / (cm[0, 0] + cm[0, 1])
            precision = cm[1, 1] / (cm[0, 1] + cm[1, 1])
            recall = val_metrics["sensitivity"]
            val_metrics["F1 score"] = 2 * (precision * recall) / (precision + recall)
            logging.info(
                "Epoch {:d}/{:d} - [Youden: {:.8f}] Accuracy: {:.3f} | Sensitivity: {:.3f} | Specificity: {:.3f} | F1 Score: {:.3f}"
                .format(cur_epoch, config.TOTAL_EPOCH, youden_thresh,
                        val_metrics["accuracy"], val_metrics["sensitivity"], val_metrics["specificity"], val_metrics["F1 score"]))

        # ----------------------------------------
        #               Save Model
        # ----------------------------------------
        if cur_epoch >= config.SAVE_START:
            if cur_epoch % config.SAVE_INTERVAL == 0 or cur_epoch == config.TOTAL_EPOCH:
                save_model(netC, 'cls_{:d}.pth'.format(cur_epoch), config.CKPT_PATH)
                logging.info('The trained classifier is successfully saved at epoch {:d}\n'.format(cur_epoch))

    print("Training finished !!")


def test_classifier(config):

    # saving path
    result_path = os.path.join(config.TEST_PATH, str(config.MODEL_CLS))
    os.makedirs(result_path, exist_ok=True)
    # ----------------------------------------
    #       Initialize testing dataset
    # ----------------------------------------

    # Define testing dataset
    test_data = PETDataset(config.DATAROOT, config.VAL_DATA, config.IMAGE_SIZE, config.DATA_NORM)
    print('The overall number of testing images is %d' % len(test_data))

    # Define testing dataloader
    test_loader = DataLoader(test_data, batch_size=1, num_workers=config.NUM_WORKERS)

    # ----------------------------------------
    #      Initialize testing parameters
    # ----------------------------------------

    # build classifier
    if config.MODE == 'test' and config.MODEL_CLS:
        netC_name = os.path.join(config.CKPT_PATH, f'cls_{config.MODEL_CLS}.pth')
        netC = create_classifier(config, netC_name).cuda()
        netC.eval()
        print("Load Classifier from epoch {:d}".format(config.MODEL_CLS))
    else:
        raise ValueError("config.MODEL_CLS is None! No classifier is loaded!")

    if config.IN_CHANNELS == 1:
        print(f"Classifier Input: {config.INPUT_TYPE}")
    elif config.IN_CHANNELS == 2:
        print("Classifier Input: real FDG + syn CFT")
    else:
        raise ValueError(f"{config.IN_CHANNELS} is not valid input channels")

    # build generators
    netG = create_generator(config, config.GEN_PATH).cuda()
    netG.eval()
    print("Load Generator from '{:s}'".format(config.GEN_PATH))


    # ----------------------------------------
    #                Metrics
    # ----------------------------------------

    # classification metrics
    cls_metrics = {
        "img_name": [], "paired": [],                                                        # image information
        "y_true": [], "y_score": [], "auc": -1,                                              # labels and scores
        "y_pred": [], "accuracy": -1, "sensitivity": -1, "specificity": -1, "F1 score": -1,  # Youden's index threshold
    }

    # ----------------------------------------
    #              Start Testing
    # ----------------------------------------

    print("Start testing ......")
    with torch.no_grad():
        for i, data in tqdm(enumerate(test_loader), total=len(test_loader), desc="Testing"):

            # load images ([B, 1, D, H, W]) and class label ([B, CLASS_NUM])
            real_FDG_img = data["FDG_img"].cuda().float()
            label = data["label"].cuda().float()

            save_img_name = data["img_name"][0]
            is_paired = data["paired_flag"][0]
            paired_flag = "Yes" if is_paired else "No"
            cls_metrics["img_name"].append(save_img_name)
            cls_metrics["paired"].append(paired_flag)

            # construct inputs to classifier
            if config.IN_CHANNELS > 1:
                with torch.no_grad():
                    input_CFT_img = netG(real_FDG_img)
                cls_inputs = torch.cat([real_FDG_img, input_CFT_img], dim=1)
            else:
                if config.INPUT_TYPE == "real_FDG":
                    cls_inputs = real_FDG_img
                elif config.INPUT_TYPE == "syn_CFT":
                    with torch.no_grad():
                        input_CFT_img = netG(real_FDG_img)
                    cls_inputs = input_CFT_img
                else:
                    raise ValueError(f"{config.INPUT_TYPE} is not valid input type")

            # classification results
            cls_out = netC(cls_inputs)
            cls_pred_scores = torch.softmax(cls_out, dim=1)
            cls_metrics["y_true"].append(label[:, 1].item())
            cls_metrics["y_score"].append(cls_pred_scores[:, 1].item())

    # save raw results
    np.save(os.path.join(result_path, "y_true.npy"), cls_metrics["y_true"])
    np.save(os.path.join(result_path, "y_score.npy"), cls_metrics["y_score"])

    # draw ROC curve
    fpr, tpr, thresholds = roc_curve(cls_metrics["y_true"], cls_metrics["y_score"])
    cls_metrics["auc"] = auc(fpr, tpr)
    print('AUC: {:.3f}'.format(cls_metrics["auc"]))

    plt.figure()
    plt.plot(fpr, tpr, color='darkorange', lw=2, label='ROC curve (area = %0.4f)' % cls_metrics["auc"])
    plt.plot([0, 1], [0, 1], color='navy', lw=2, linestyle='--')
    plt.xlim([-0.05, 1.0])
    plt.ylim([0.0, 1.05])
    plt.xlabel("False Positive Rate")
    plt.ylabel("True Positive Rate")
    plt.title(f"ROC Curve")
    plt.legend(loc="lower right")
    plt.savefig(os.path.join(result_path, "roc_curve.png"), format='png', dpi=300)

    # calculate classification metrics
    youden_thresh = find_youden_index_threshold(fpr, tpr, thresholds)
    cls_metrics["y_pred"] = np.where(np.asarray(cls_metrics["y_score"]) >= youden_thresh, 1, 0)
    cm = confusion_matrix(cls_metrics["y_true"], cls_metrics["y_pred"])
    cls_metrics["accuracy"] = accuracy_score(cls_metrics["y_true"], cls_metrics["y_pred"])
    cls_metrics["sensitivity"] = cm[1, 1] / (cm[1, 0] + cm[1, 1])
    cls_metrics["specificity"] = cm[0, 0] / (cm[0, 0] + cm[0, 1])
    precision = cm[1, 1] / (cm[0, 1] + cm[1, 1])
    recall = cls_metrics["sensitivity"]
    cls_metrics["F1 score"] = 2 * (precision * recall) / (precision + recall)
    print(
        '[Youden: {:.8f}] Accuracy: {:.3f} | Sensitivity: {:.3f} | Specificity: {:.3f} | F1 Score: {:.3f}'
        .format(youden_thresh,
                cls_metrics["accuracy"], cls_metrics["sensitivity"], cls_metrics["specificity"], cls_metrics["F1 score"])
    )

    disp = ConfusionMatrixDisplay(confusion_matrix=cm, display_labels=["NC", "PD"])
    disp.plot(cmap="Blues")
    plt.title("Confusion Matrix (Score Threshold {:.3f})".format(youden_thresh))
    plt.savefig(os.path.join(result_path, f"confusion_matrix.png"))

    # record results
    # (1) overall result to txt
    record_file = os.path.join(result_path, "overall_result.txt")
    with open(record_file, "w", encoding='utf-8') as f:
        f.write(
            "AUC: {:.3f}\n"
            "[Youden: {:.8f}] Accuracy: {:.3f} | Sensitivity: {:.3f} | Specificity: {:.3f} | F1 Score: {:.3f}\n"
            .format(cls_metrics["auc"], youden_thresh, cls_metrics["accuracy"],
                    cls_metrics["sensitivity"], cls_metrics["specificity"], cls_metrics["F1 score"])
        )

    # (2) every sample results to excel
    every_sample_df = pd.DataFrame({"Image Name": cls_metrics["img_name"]}, index=range(1, len(cls_metrics["img_name"])+1))
    every_sample_df["Paired"] = cls_metrics["paired"]
    every_sample_df["Predicted Score"] = cls_metrics["y_score"]
    every_sample_df["Prediction (Youden Thresh)"] = [
        "PD" if pred == 1 else "NC" for pred in cls_metrics["y_pred"]
    ]

    xls_file = os.path.join(result_path, 'every_sample_results.xlsx')
    with pd.ExcelWriter(xls_file, engine='openpyxl') as writer:
        every_sample_df.to_excel(writer, index=False)
        worksheet = writer.sheets["Sheet1"]
        for row in worksheet.iter_rows(min_row=1, max_row=worksheet.max_row,
                                       min_col=1, max_col=worksheet.max_column):
            for cell in row:
                cell.number_format = '0.0000'
                if cell.row == 1:
                    cell.font = openpyxl.styles.Font(name='Times New Roman', size=14)
                else:
                    cell.font = openpyxl.styles.Font(name='Times New Roman', size=12)

    print("Testing finished !!")
